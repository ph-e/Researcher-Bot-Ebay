import openpyxl
import requests
import re
import cloudinary.uploader
import os
import random
import string
from bs4 import BeautifulSoup

def parse(product_url):
    #We receive the title and links to product photos on ebay
    dataarray = []
    response = requests.get(product_url)
    if response.status_code == 200:
        text = response.text
        pattern = r'"title":"(.*?) - Picture[^"]*"'
        pText = re.search(pattern, text)

        if pText:
            captured_text = pText.group(1)
            title = re.sub(r'.*?"title":"', '', captured_text)
            dataarray.append('For ' + title)
        soup = BeautifulSoup(text, 'html.parser')



        
        # Extracting price
        price_div = soup.find('div', class_='x-price-primary')
        if price_div:
            price_span = price_div.find('span', class_='ux-textspans')
    
            if price_span:
                dataarray.append('Price: ' + price_span.text)



        # Extracting shipping price
        shipping_div = soup.find('div', class_='ux-labels-values__values-content')
        if shipping_div:
            shipping = shipping_div.find('span', class_='ux-textspans ux-textspans--BOLD')
            if shipping:
                dataarray.append(shipping.text)

        # Extracting stock
        stock_div = soup.find('div', class_='d-quantity__availability')
        if stock_div:
            stock_span = stock_div.find('span', class_='ux-textspans')
            if stock_span:
                dataarray.append(stock_span.text)

        # Extrecting Delivery
        delivery_div = soup.find_all('div', class_='ux-labels-values__values-content')
        if len(delivery_div) >= 2:
            delivery_values_div = delivery_div[1]
            spans = delivery_values_div.find_all('span', class_='ux-textspans')
            delivery_value = ' '.join(span.text for span in spans[:4])
            dataarray.append(delivery_value)

        # Extrecting Seller
        seller_info_div = soup.find('div', class_='x-sellercard-atf__info__about-seller')
        if seller_info_div:
            seller_span = seller_info_div.find('span', class_='ux-textspans ux-textspans--BOLD')
            if seller_span:
                dataarray.append(seller_span.text)




        pattern = r'"(https://i\.ebayimg\.com/images/[^"]*/s-l1600\.(jpg|png))"'
        matches = re.findall(pattern, text)
        for match in matches:
            if match[0] not in dataarray:  # Проверка на уникальность
                dataarray.append(match[0])

        print('Information copied successfully')
    else:
        print(f"Failed to retrieve content. Status code: {response.status_code}")

    return dataarray

def upload_image_to_cloudinary(api_key, api_secret, cloud_name, image_path):
    #uploading downloaded photos and returning a link to them
    try:
        cloudinary.config(
            cloud_name=cloud_name,
            api_key=api_key,
            api_secret=api_secret
        )

        response = cloudinary.uploader.upload(image_path)

        if 'secure_url' in response:
            uploaded_image_url = response['secure_url']
            print("Image uploaded successfully.")
            return uploaded_image_url
        else:
            print("Image upload failed. Response:", response)
    except Exception as e:
        print("An error occurred:", str(e))

    return ""

def download_image(image_url, save_path):
    #download photos from the link
    try:
        response = requests.get(image_url)
        if response.status_code == 200:
            with open(save_path, 'wb') as file:
                file.write(response.content)
            print("Image downloaded successfully.")
        else:
            print("Failed to download the image. HTTP Status Code:", response.status_code)
    except Exception as e:
        print("An error occurred while downloading the image:", str(e))

def getLinks(urlSeller):
    #We receive links to all products located on the page
    response = requests.get(urlSeller)
    if response.status_code == 200:
        text = response.text
        pattern = r'https://www\.ebay\.com/itm/[^"\s]+'
        links = re.findall(pattern, text)
        # Удаляем строку "> <div" из каждой ссылки
        links = [re.sub(r'><div', '', link) for link in links]
        return links
    return ''

def main():

    #We get a link to the product page
    urlSeller = input("Enter the link to the seller's page:")

    #Count of images
    countIMG = int(input("Enter the maximum number of images to upload:"))

    #Open a text document for reading
    with open("config.txt", "r") as file:
        lines = file.readlines()

    #Loop through strings and look for key values
    api_key = None
    api_secret = None
    cloud_name = None

    for line in lines:
        if "api_key" in line:
            api_key = line.split("=")[1].strip()
        elif "api_secret" in line:
            api_secret = line.split("=")[1].strip()
        elif "cloud_name" in line:
            cloud_name = line.split("=")[1].strip()

    #creating a new excel table
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    #Receiving a link to the seller's items
    links = getLinks(urlSeller)
    info = []
    processed_links = set()
    for idx, link in enumerate(links):
        if link not in processed_links: # checking for duplicate links
            if link != "":
                processed_links.add(link)
                info = parse(link)
                indexRow = len(processed_links)
                worksheet.cell(row=indexRow, column=1, value=link)
                for index, item in enumerate(info):
                    if index == 0:
                        worksheet.cell(row=indexRow, column=2, value=info[0])
                    elif index == 1:
                        worksheet.cell(row=indexRow, column=3, value=item)
                    elif index == 2:
                        worksheet.cell(row=indexRow, column=4, value=item)
                    elif index == 3:
                        worksheet.cell(row=indexRow, column=5, value=item)
                    elif index == 4:
                        worksheet.cell(row=indexRow, column=6, value=item)
                    elif index == 5:
                        worksheet.cell(row=indexRow, column=7, value=item)
                    elif index <= (countIMG + 5):
                        image_url = item
                        image_path = f"downloaded_images/{indexRow}.jpg"
                        try:
                            if image_url:
                                download_image(image_url, image_path)
                                urlIMG = upload_image_to_cloudinary(api_key, api_secret, cloud_name, image_path)
                                os.remove(image_path)
                                worksheet.cell(row=indexRow, column=index+2, value=urlIMG)
                        except Exception as e:
                            print(f"ERROR Failed to download image: {e}")

    #Generating a random file name
    file_name = ''.join(random.choice(string.ascii_letters + string.digits) for i in range(8))                    
    # Save the modified workbook
    workbook.save(file_name + ".xlsx")

    # Close the workbook
    workbook.close()
    message = f"Your page has been successfully processed and saved in the file {file_name}.\nPlease check the file to ensure the code is executed correctly.\nYou can close this window by pressing any button!"
    input(message)

if __name__ == "__main__":
    main()
